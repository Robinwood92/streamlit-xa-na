import streamlit as st
import geopandas as gpd
import folium
from folium.plugins import Draw, FloatImage
from shapely.geometry import shape
from shapely.ops import unary_union
from streamlit_folium import st_folium
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import requests
from datetime import datetime, timedelta, timezone
import os
import base64
from PIL import Image
import io
import asyncio
import sys
if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
# =====================
# ⚙️ CẤU HÌNH CROP RADAR
# =====================
RADAR_URL = "https://iweather.gov.vn/dashboard/?productRadar=CMAX&areaRadar=VIN"

# Vùng cần crop (tọa độ địa lý)
CROP_MIN_LAT = 18.3
CROP_MAX_LAT = 20.5
CROP_MIN_LON = 103.5
CROP_MAX_LON = 106.1

# ⚠️ Vị trí chèn ảnh vào Excel — thay đổi khi bạn xác định ô cụ thể
RADAR_IMG_CELL = "B14"  # Ảnh sẽ được neo tại B14, kéo dài đến G23

# =====================
# 🛰️ HÀM LẤY URL ẢNH RADAR
# =====================
def get_vin_radar_urls():
    """Lấy 2 URL ảnh radar mới nhất từ trạm VIN"""
    base_url = "http://hymetnet.gov.vn/dataout_web/VIN"
    
    t1 = datetime.now(timezone.utc) - timedelta(minutes=10)
    t1 = t1.replace(minute=(t1.minute // 10) * 10, second=0, microsecond=0)
    
    t0 = t1 - timedelta(minutes=10)
    t0 = t0.replace(minute=(t0.minute // 10) * 10, second=0, microsecond=0)
    
    def fmt(dt):
        ymd = dt.strftime("%Y%m%d")
        ymdhm = dt.strftime("%Y%m%d%H%M")
        dt_utc7 = dt + timedelta(hours=7)
        display_time = dt_utc7.strftime("%H:%M")
        return ymd, ymdhm, display_time, dt
    
    ymd0, ymdhm0, display0, dt0 = fmt(t0)
    ymd1, ymdhm1, display1, dt1 = fmt(t1)
    
    url0 = f"{base_url}/{ymd0}/VIN_{ymdhm0}_CMAX00.png"
    url1 = f"{base_url}/{ymd1}/VIN_{ymdhm1}_CMAX00.png"
    
    return [(ymdhm0, url0, display0, dt0), (ymdhm1, url1, display1, dt1)]

# =====================
# 🌧️ TẢI VÀ CHUYỂN ẢNH RADAR SANG BASE64
# =====================
def download_radar_as_base64(url):
    try:
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            img_base64 = base64.b64encode(r.content).decode()
            return f"data:image/png;base64,{img_base64}"
        return None
    except:
        return None

def load_all_radars():
    urls = get_vin_radar_urls()
    loaded_radars = []
    for timecode, url, display_time, dt in urls:
        try:
            radar_base64 = download_radar_as_base64(url)
            if radar_base64:
                loaded_radars.append((timecode, radar_base64, display_time, dt))
        except:
            pass
    return loaded_radars

# =====================
# 🎨 TẢI LEGEND RADAR
# =====================
@st.cache_data
def load_legend_base64():
    try:
        with open("legend_radar.jpg", "rb") as f:
            legend_data = f.read()
        legend_base64 = base64.b64encode(legend_data).decode()
        return f"data:image/jpg;base64,{legend_base64}"
    except FileNotFoundError:
        st.warning("⚠️ Không tìm thấy file 'legend_radar.jpg' trong cùng thư mục.")
        return None

# =====================
# ⚙️ Tải shapefile Nghệ An
# =====================
@st.cache_data
def load_shapefile():
    gdf = gpd.read_file("Xa_NA_chuan.geojson")
    return gdf.to_crs(epsg=4326)

gdf = load_shapefile()

# =====================
# 📝 HÀM GHI VÀO MERGED CELL
# =====================
def write_to_merged_cell(ws, row, col, value):
    """Ghi giá trị vào ô có thể là merged cell: unmerge → ghi → re-merge"""
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            merge_str = str(merged_range)
            ws.unmerge_cells(merge_str)
            ws.cell(row=merged_range.min_row,
                    column=merged_range.min_col, value=value)
            ws.merge_cells(merge_str)
            return
    ws.cell(row=row, column=col, value=value)

# =====================
# 🧹 HÀM XÓA DỮ LIỆU TỪ DÒNG 46 ĐẾN DÒNG CUỐI CÓ DỮ LIỆU
# =====================
def clear_rows_from_46(ws):
    last_row = 46
    for row in ws.iter_rows():
        for cell in row:
            if cell.row >= 46 and cell.value is not None:
                if cell.row > last_row:
                    last_row = cell.row
    for row_idx in range(46, last_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            try:
                ws.cell(row=row_idx, column=col_idx).value = None
            except AttributeError:
                pass

# =====================
# 🙈 HÀM ẨN DÒNG TỪ 60 ĐẾN DÒNG CUỐI CÓ DỮ LIỆU
# =====================
def hide_rows_60_to_last(ws):
    last_row = 59
    for row in ws.iter_rows():
        for cell in row:
            if cell.row >= 60 and cell.value is not None:
                if cell.row > last_row:
                    last_row = cell.row
    if last_row >= 60:
        for row_idx in range(60, last_row + 1):
            ws.row_dimensions[row_idx].hidden = True

# =====================
# 📸 HÀM CHỤP VÀ CROP ẢNH RADAR TỪ WEBSITE
# =====================
async def _capture_radar_async():
    """Dùng Playwright chụp trang radar, tự động cài đặt trình duyệt nếu thiếu"""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return None, "❌ Chưa cài Playwright trong requirements.txt"

    try:
        # Tự động cài đặt chromium nếu chưa có (dành cho Streamlit Cloud)
        import subprocess
        subprocess.run(["playwright", "install", "chromium"], check=False)

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",          # Bắt buộc trên Linux Container
                    "--disable-gpu",         # Giúp ổn định hơn trên server
                    "--disable-dev-shm-usage" # Tránh lỗi bộ nhớ đệm trên Docker
                ]
            )

            context = await browser.new_context(
                viewport={"width": 1920, "height": 1080},
                device_scale_factor=3,   # ↑ tăng lên 3x → ảnh sắc nét hơn nhiều
                user_agent="Mozilla/5.0"
            )

            page = await context.new_page()

            # ✅ giữ URL cũ của bạn (không đổi)
            await page.goto(RADAR_URL, wait_until="domcontentloaded", timeout=60000)

            # ✅ chờ map load đúng cách (tránh timeout)
            await page.wait_for_selector("canvas", timeout=20000)
            await page.wait_for_timeout(4000)

            # ✅ Chờ Leaflet map sẵn sàng
            for _ in range(20):
                map_el = await page.query_selector(".leaflet-container")
                if map_el:
                    break
                await page.wait_for_timeout(1000)
            else:
                return None, "❌ Không tìm thấy bản đồ sau 20 giây"

            # ✅ Zoom IN Leaflet vào tâm vùng cần chụp (không dùng fitBounds để tránh zoom out)
            center_lat = (CROP_MIN_LAT + CROP_MAX_LAT) / 2
            center_lon = (CROP_MIN_LON + CROP_MAX_LON) / 2
            zoom_js = f"""() => {{
                try {{
                    let map = null;
                    // Cách 1: biến toàn cục phổ biến
                    if (window._map && window._map.setView) map = window._map;
                    else if (window.map && window.map.setView) map = window.map;
                    // Cách 2: duyệt window
                    if (!map) {{
                        for (const key of Object.keys(window)) {{
                            const obj = window[key];
                            if (obj && typeof obj === 'object' && obj.fitBounds && obj.setView) {{
                                map = obj; break;
                            }}
                        }}
                    }}
                    // Cách 3: lấy từ DOM element nội bộ Leaflet
                    if (!map) {{
                        const el = document.querySelector('.leaflet-container');
                        if (el && el._leaflet_map) map = el._leaflet_map;
                    }}
                    if (!map) return map.getZoom();
                    const currentZoom = map.getZoom();
                    const targetZoom = 8;  // zoom level cho vùng Nghệ An (~2° × 2.6°)
                    // Chỉ zoom IN — không bao giờ zoom out
                    const newZoom = Math.max(currentZoom, targetZoom);
                    map.setView([{center_lat}, {center_lon}], newZoom, {{ animate: false }});
                    return newZoom;
                }} catch(e) {{ return -1; }}
            }}"""

            zoom_result = await page.evaluate(zoom_js)
            if zoom_result and zoom_result > 0:
                await page.wait_for_timeout(2000)  # chờ tiles reload sau setView
            else:
                await page.wait_for_timeout(500)

            # =====================
            # 🖱️ Focus vào map + scroll zoom IN thêm (theo gợi ý ChatGPT, cải tiến)
            # =====================
            box = await map_el.bounding_box()
            cx = box["x"] + box["width"]  / 2
            cy = box["y"] + box["height"] / 2

            # Click vào giữa map để đảm bảo nhận sự kiện wheel
            await page.mouse.click(cx, cy)

            # Scroll zoom IN 2 lần (scroll up = zoom in trong Leaflet)
            SCROLL_TIMES = 2        # tăng nếu muốn zoom sâu hơn
            SCROLL_DELTA = -300     # âm = lên = zoom in; giảm xuống -600 nếu mỗi bước zoom ít
            for _ in range(SCROLL_TIMES):
                await page.mouse.wheel(0, SCROLL_DELTA)
                await page.wait_for_timeout(1500)   # chờ tile load sau mỗi bước

            # =====================
            # 📸 Chụp full page rồi crop theo bounding box của map element
            # =====================
            full_img_bytes = await page.screenshot(
                full_page=False,    # chụp viewport (không scroll thêm)
                type="png",
                scale="device"      # device pixel ratio → sắc nét
            )
            await browser.close()

            # Crop chính xác vùng map element từ ảnh full viewport
            img  = Image.open(io.BytesIO(full_img_bytes))
            dpr  = 3  # phải khớp với device_scale_factor ở trên
            x1   = int(box["x"]      * dpr)
            y1   = int(box["y"]      * dpr)
            x2   = int((box["x"] + box["width"])  * dpr)
            y2   = int((box["y"] + box["height"]) * dpr)
            cropped = img.crop((x1, y1, x2, y2))

            buf = io.BytesIO()
            cropped.save(buf, format='PNG')
            buf.seek(0)

            return buf, None

    except Exception as e:
        import traceback
        return None, f"❌ Lỗi Playwright: {type(e).__name__}: {str(e)}\n\n{traceback.format_exc()}"


def capture_radar_crop():
    """Wrapper đồng bộ cho hàm async"""
    return asyncio.run(_capture_radar_async())


# =====================
# 🧭 Giao diện
# =====================
st.set_page_config(layout="wide")
st.title("🗺️ Bản đồ Nghệ An với Radar và chọn vùng")

# =====================
# 📡 Sidebar - Cài đặt Radar
# =====================
with st.sidebar:
    st.header("📡 Cài đặt lớp Radar")
    show_radar = st.checkbox("Hiển thị ảnh Radar", value=True)
    
    if show_radar:
        radar_opacity = st.slider("Độ trong suốt Radar", 0.0, 1.0, 0.6, 0.1)
        
        with st.spinner("Đang tải ảnh radar..."):
            loaded_radars = load_all_radars()
        
        if loaded_radars:
            st.success(f"✅ Đã tải {len(loaded_radars)} ảnh radar")
            
            if len(loaded_radars) > 1:
                radar_idx = st.slider(
                    "Chọn thời điểm radar:",
                    0,
                    len(loaded_radars) - 1,
                    len(loaded_radars) - 1,
                    format=""
                )
                timecode, radar_base64, display_time, dt = loaded_radars[radar_idx]
                st.info(f"🕐 {display_time} (UTC+7)")
            else:
                timecode, radar_base64, display_time, dt = loaded_radars[0]
                st.info(f"🕐 {display_time} (UTC+7)")
        else:
            st.warning("⚠️ Không tìm thấy ảnh radar khả dụng")
            show_radar = False

    st.divider()

    # =====================
    # 📸 Nút chụp ảnh Radar từ website
    # =====================
    st.header("📸 Chụp ảnh Radar web")
    st.caption(f"Vùng crop: {CROP_MIN_LAT}–{CROP_MAX_LAT}°N, {CROP_MIN_LON}–{CROP_MAX_LON}°E")

    if st.button("📸 Chụp màn hình Radar", use_container_width=True):
        with st.spinner("Đang chụp và crop ảnh radar..."):
            img_buf, err = capture_radar_crop()
        if err:
            st.error(err)
        else:
            st.session_state.radar_screenshot = img_buf
            st.success("✅ Đã chụp xong!")

    if "radar_screenshot" in st.session_state and st.session_state.radar_screenshot:
        st.image(st.session_state.radar_screenshot, caption="Preview ảnh radar đã crop", use_container_width=True)
        if st.button("🗑️ Xóa ảnh", use_container_width=True):
            st.session_state.radar_screenshot = None
            st.rerun()

# =====================
# 🗺️ Bản đồ nền
# =====================
center = [19.23, 104.8]
m = folium.Map(location=center, zoom_start=9, tiles="OpenStreetMap")

folium.GeoJson(
    gdf,
    name="📍 Các xã Nghệ An",
    style_function=lambda x: {"color": "gray", "weight": 1, "fillOpacity": 0.1},
    tooltip=folium.GeoJsonTooltip(fields=["Xa", "Diem"], aliases=["Xã:", "Huyện:"]),
).add_to(m)

# =====================
# 🛰️ Thêm lớp Radar vào bản đồ
# =====================
if show_radar and loaded_radars:
    center_lat, center_lon = 18.656, 105.71083
    radius_deg = 2.8

    bounds = [
        [center_lat - radius_deg, center_lon - radius_deg],
        [center_lat + radius_deg, center_lon + radius_deg]
    ]

    folium.raster_layers.ImageOverlay(
        image=radar_base64,
        bounds=bounds,
        opacity=radar_opacity,
        name=f"🌧️ Radar {display_time}",
        interactive=False,
        cross_origin=False,
        zindex=1
    ).add_to(m)

    legend_base64 = load_legend_base64()
    if legend_base64:
        legend_html = f'''
        <div style="
            position: fixed;
            bottom: 25px;
            left: 10px;
            width: 42px;
            height: auto;
            z-index: 9999;
            background-color: rgba(255, 255, 255, 0.9);
            border: 2px solid grey;
            border-radius: 5px;
            padding: 5px;
        ">
            <img src="{legend_base64}" style="width: 100%; height: auto;">
        </div>
        '''
        m.get_root().html.add_child(folium.Element(legend_html))

# =====================
# ✏️ Công cụ vẽ
# =====================
Draw(
    export=False,
    draw_options={
        "polygon": {"allowIntersection": False, "showArea": True, "repeatMode": True},
        "rectangle": False,
        "circle": False,
        "circlemarker": False,
        "polyline": False,
        "marker": False,
    },
    edit_options={"edit": True, "remove": True},
).add_to(m)

folium.LayerControl(collapsed=False).add_to(m)

# =====================
# 📝 Hướng dẫn
# =====================
st.markdown("### ✏️ Hướng dẫn:")
st.markdown("""
- ☑️ Bật/tắt lớp **Radar** trong sidebar bên trái
- 🎨 Điều chỉnh độ trong suốt của ảnh radar
- 📸 Nhấn **Chụp màn hình Radar** để lấy ảnh từ website (sẽ được chèn vào Excel)
- ✏️ Dùng công cụ **Polygon** để vẽ vùng (double-click để kết thúc)
- 📍 Có thể vẽ **nhiều vùng**
- 🔄 Khi hoàn tất, nhấn **[Lấy xã]** để liệt kê các xã trong tất cả vùng đã vẽ
""")

# =====================
# 📍 Hiển thị bản đồ
# =====================
map_data = st_folium(m, height=600, width=950, returned_objects=["all_drawings"])

# =====================
# 💾 Lưu các polygon
# =====================
if "all_polygons" not in st.session_state:
    st.session_state.all_polygons = []

if map_data and "all_drawings" in map_data and map_data["all_drawings"]:
    st.session_state.all_polygons = [
        shape(feature["geometry"]) for feature in map_data["all_drawings"]
        if feature.get("geometry") is not None
    ]

st.info(f"📍 Hiện có **{len(st.session_state.all_polygons)}** vùng được vẽ.")

# =====================
# 🚀 Nút LẤY XÃ
# =====================
if st.button("📍 Lấy xã trong tất cả vùng đã vẽ"):
    if st.session_state.all_polygons:
        try:
            union_polygon = unary_union(st.session_state.all_polygons)
            selected_gdf = gdf[gdf.intersects(union_polygon)]

            if not selected_gdf.empty:
                st.success(f"✅ Tìm thấy {len(selected_gdf)} xã nằm trong các vùng đã vẽ.")

                grouped_df = (
                    selected_gdf.groupby("Diem")["Xa"]
                    .apply(lambda x: ", ".join(sorted(set(x))))
                    .reset_index()
                )

                st.markdown("## 🗂️ Danh sách xã theo huyện")
                for diem, xa_list in grouped_df.values:
                    st.write(f"**{diem}**: {xa_list}")

                # --- Ghi dữ liệu vào file template.xlsx
                try:
                    wb = load_workbook("template.xlsx")
                    ws = wb.active

                    # Bước 1: Lưu max_row của template trước khi xóa
                    template_max_row = ws.max_row

                    # Bước 2: Xóa dữ liệu cũ từ dòng 46
                    clear_rows_from_46(ws)

                    # Bước 3: Ghi dữ liệu mới từ dòng 46
                    start_row = 46
                    for i, row in enumerate(grouped_df.itertuples(index=False), start=start_row):
                        write_to_merged_cell(ws, i, 1, row.Diem)
                        write_to_merged_cell(ws, i, 3, row.Xa)

                    # Bước 4: Ẩn/hiện dòng từ 60 đến template_max_row
                    # Dòng đã ghi (≤ last_written_row): hiện — dòng chưa ghi: ẩn
                    last_written_row = start_row + len(grouped_df) - 1
                    for row_idx in range(60, template_max_row + 1):
                        ws.row_dimensions[row_idx].hidden = (row_idx > last_written_row)

                    # Bước 4: Chèn ảnh radar vào Excel (nếu đã chụp)
                    radar_buf = st.session_state.get("radar_screenshot")
                    if radar_buf:
                        try:
                            radar_buf.seek(0)

                            # Tính kích thước pixel khớp với vùng B14:G23
                            from openpyxl.utils import column_index_from_string, get_column_letter

                            START_COL, START_ROW = "B", 14
                            END_COL,   END_ROW   = "F", 23

                            def col_width_px(ws, col_letter):
                                cd = ws.column_dimensions.get(col_letter)
                                w = cd.width if cd and cd.width else 8.43
                                return int(w * 7 + 5)

                            def row_height_px(ws, row_num):
                                rd = ws.row_dimensions.get(row_num)
                                h = rd.height if rd and rd.height else 15
                                return int(h * 96 / 72)

                            sc = column_index_from_string(START_COL)
                            ec = column_index_from_string(END_COL)
                            total_w = sum(col_width_px(ws, get_column_letter(c)) for c in range(sc, ec + 1))
                            total_h = sum(row_height_px(ws, r) for r in range(START_ROW, END_ROW + 1))

                            # ✅ KHÔNG resize PIL → giữ nguyên độ phân giải gốc
                            # openpyxl tự scale ảnh khớp vùng B14:F23
                            radar_buf.seek(0)
                            xl_img = XLImage(radar_buf)
                            xl_img.anchor = RADAR_IMG_CELL
                            xl_img.width  = total_w   # pixel width của vùng Excel
                            xl_img.height = total_h   # pixel height của vùng Excel
                            ws.add_image(xl_img)
                            st.info(f"🖼️ Đã chèn ảnh radar vào vùng **B14:F23** ({total_w}×{total_h}px)")
                        except Exception as img_err:
                            st.warning(f"⚠️ Không thể chèn ảnh radar: {img_err}")
                    else:
                        st.info("ℹ️ Chưa có ảnh radar — nhấn **Chụp màn hình Radar** ở sidebar để thêm vào Excel.")

                    output = BytesIO()
                    now = datetime.now()
                    filename_base = now.strftime("NGAN_DONG_%Y%m%d_%H%M")
                    excel_filename = f"{filename_base}.xlsx"
                    wb.save(excel_filename)
                    output.seek(0)

                    st.download_button(
                        label="📥 Tải file Excel (theo template)",
                        data=output,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except FileNotFoundError:
                    st.error("❌ Không tìm thấy file 'template.xlsx' trong cùng thư mục.")
            else:
                st.warning("⚠️ Không có xã nào nằm trong các vùng đã vẽ.")
        except Exception as e:
            st.error(f"❌ Lỗi xử lý vùng vẽ: {e}")
    else:
        st.warning("⚠️ Bạn chưa vẽ vùng nào trên bản đồ.")
else:
    st.info("🖱️ Hãy vẽ vùng rồi nhấn **Lấy xã** để bắt đầu.")
